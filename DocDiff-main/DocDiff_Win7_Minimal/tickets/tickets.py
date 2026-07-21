from __future__ import annotations

import csv
import json
import os
import re
from dataclasses import asdict, dataclass
from typing import Any, Dict, List, Mapping, Optional, Sequence, Union


@dataclass
class Ticket:
    """人工问题单条目。

    序号: 与更改单「问题N」对齐（默认 N 从 1 起，可与 --problem-start 对齐）
    问题: 问题描述，如「xxx需求变更」
    问题单编号: 一次版本变更内唯一，形如 项目型号-WT-两位序号
              例：DFKS112-WT-01、DFKS112-WT-02（DFKS112=项目型号）
    """

    seq: int
    title: str = ""
    ticket_no: str = ""

    def display_no(self) -> str:
        return (self.ticket_no or "").strip()

    def display_title(self) -> str:
        return (self.title or "").strip()


# 默认流水号宽度：DFKS112-WT-01
DEFAULT_TICKET_SEQ_WIDTH = 2


_HEADER_ALIASES = {
    "seq": {"序号", "seq", "index", "问题序号", "no", "编号序号"},
    "title": {"问题", "title", "描述", "问题描述", "说明", "summary"},
    "ticket_no": {
        "问题单编号",
        "ticket_no",
        "ticket",
        "ticket_id",
        "单号",
        "问题单号",
    },
}


def _norm_header(h: str) -> str:
    return re.sub(r"\s+", "", (h or "").strip().lower())


def _map_headers(headers: Sequence[str]) -> Dict[str, int]:
    """返回逻辑字段 -> 列下标。优先精确匹配，避免「问题」误匹配「问题单编号」。"""
    out: Dict[str, int] = {}
    norms = [(idx, _norm_header(str(raw)), str(raw).strip()) for idx, raw in enumerate(headers)]

    # 第一遍：精确匹配（含去空白后完全相等）
    for idx, key, _raw in norms:
        if not key:
            continue
        for field, aliases in _HEADER_ALIASES.items():
            if field in out:
                continue
            for a in aliases:
                if key == _norm_header(a):
                    out[field] = idx
                    break

    # 第二遍：仅对仍未命中的字段做「表头以别名结尾」的宽松匹配（如 外部问题单编号）
    for idx, key, _raw in norms:
        if not key:
            continue
        if idx in out.values():
            continue
        for field, aliases in _HEADER_ALIASES.items():
            if field in out:
                continue
            for a in aliases:
                na = _norm_header(a)
                if not na or len(na) < 2:
                    continue
                # 表头以别名结尾，且别名长度足够，避免「问题」命中「问题单编号」
                if key.endswith(na) and len(na) >= 4:
                    out[field] = idx
                    break
                if na.endswith(key) and len(key) >= 4:
                    out[field] = idx
                    break

    # 位置回退：第0列序号 第1列问题 第2列单号
    if "seq" not in out and len(headers) >= 1:
        out["seq"] = 0
    if "title" not in out and len(headers) >= 2:
        out["title"] = 1
    if "ticket_no" not in out and len(headers) >= 3:
        out["ticket_no"] = 2
    return out


def _parse_seq(val: Any) -> Optional[int]:
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    # 允许「1」「1.」「问题1」
    m = re.search(r"(\d+)", s)
    if not m:
        return None
    try:
        return int(m.group(1))
    except ValueError:
        return None


def _row_to_ticket(row: Sequence[Any], colmap: Mapping[str, int]) -> Optional[Ticket]:
    def cell(field: str) -> str:
        i = colmap.get(field)
        if i is None or i >= len(row):
            return ""
        v = row[i]
        return "" if v is None else str(v).strip()

    seq = _parse_seq(cell("seq") if "seq" in colmap else (row[0] if row else ""))
    if seq is None:
        return None
    title = cell("title")
    ticket_no = cell("ticket_no")
    if not title and not ticket_no:
        return None
    return Ticket(seq=seq, title=title, ticket_no=ticket_no)


def _load_json(path: str) -> List[Ticket]:
    with open(path, "r", encoding="utf-8-sig") as f:
        data = json.load(f)
    items: List[Any]
    if isinstance(data, dict):
        items = data.get("tickets") or data.get("items") or data.get("问题单") or []
    elif isinstance(data, list):
        items = data
    else:
        raise ValueError("问题单 JSON 须为数组，或含 tickets/items 字段的对象")

    tickets: List[Ticket] = []
    for it in items:
        if not isinstance(it, dict):
            continue
        seq = _parse_seq(
            it.get("seq", it.get("序号", it.get("index", it.get("问题序号"))))
        )
        if seq is None:
            continue
        title = str(
            it.get("title")
            or it.get("问题")
            or it.get("描述")
            or it.get("问题描述")
            or ""
        ).strip()
        ticket_no = str(
            it.get("ticket_no")
            or it.get("问题单编号")
            or it.get("ticket_id")
            or it.get("单号")
            or it.get("编号")
            or ""
        ).strip()
        if not title and not ticket_no:
            continue
        tickets.append(Ticket(seq=seq, title=title, ticket_no=ticket_no))
    return tickets


def _load_csv(path: str) -> List[Ticket]:
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        # 尝试检测分隔符
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
        except csv.Error:
            dialect = csv.excel
        reader = csv.reader(f, dialect)
        rows = list(reader)
    if not rows:
        return []
    # 是否有表头
    first = [str(c).strip() for c in rows[0]]
    has_header = any(
        _norm_header(c) in {_norm_header(a) for aliases in _HEADER_ALIASES.values() for a in aliases}
        for c in first
    )
    if has_header:
        colmap = _map_headers(first)
        data_rows = rows[1:]
    else:
        colmap = {"seq": 0, "title": 1, "ticket_no": 2}
        data_rows = rows

    tickets: List[Ticket] = []
    for row in data_rows:
        if not row or all(not str(c).strip() for c in row):
            continue
        t = _row_to_ticket(row, colmap)
        if t:
            tickets.append(t)
    return tickets


def _load_xlsx(path: str) -> List[Ticket]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:
        raise ImportError(
            "读取 .xlsx 问题单需要 openpyxl。请 pip install openpyxl，"
            "或改用 .csv / .json 台账。"
        ) from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    if not rows:
        return []
    first = ["" if c is None else str(c).strip() for c in rows[0]]
    has_header = any(
        _norm_header(c) in {_norm_header(a) for aliases in _HEADER_ALIASES.values() for a in aliases}
        for c in first
        if c
    )
    if has_header:
        colmap = _map_headers(first)
        data_rows = rows[1:]
    else:
        colmap = {"seq": 0, "title": 1, "ticket_no": 2}
        data_rows = rows

    tickets: List[Ticket] = []
    for row in data_rows:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        t = _row_to_ticket(list(row), colmap)
        if t:
            tickets.append(t)
    return tickets


def load_tickets(path: str) -> Dict[int, Ticket]:
    """加载问题单台账，返回 {序号: Ticket}。后出现的同序号覆盖先前。"""
    if not path or not str(path).strip():
        return {}
    path = os.path.abspath(path)
    if not os.path.isfile(path):
        raise FileNotFoundError(f"问题单文件不存在：{path}")

    ext = os.path.splitext(path)[1].lower()
    if ext == ".json":
        items = _load_json(path)
    elif ext in {".csv", ".tsv", ".txt"}:
        items = _load_csv(path)
    elif ext in {".xlsx", ".xlsm"}:
        items = _load_xlsx(path)
    else:
        # 尝试 JSON，再 CSV
        try:
            items = _load_json(path)
        except Exception:
            items = _load_csv(path)

    by_seq: Dict[int, Ticket] = {}
    for t in items:
        by_seq[int(t.seq)] = t
    return by_seq


def tickets_to_list(tickets: Mapping[int, Ticket]) -> List[dict]:
    return [asdict(tickets[k]) for k in sorted(tickets.keys())]


def normalize_ticket_prefix(prefix: str) -> str:
    """规范化前缀：去空白、去掉末尾多余连字符。

    接受：
      DFKS112-WT
      DFKS112-WT-
      DFKS112（会补全为 DFKS112-WT）
    """
    p = (prefix or "").strip().upper()
    if not p:
        return ""
    p = re.sub(r"\s+", "", p)
    p = p.rstrip("-_")
    # 仅型号时默认加 -WT
    if p and not re.search(r"-WT$", p, flags=re.IGNORECASE):
        if re.fullmatch(r"[A-Z0-9]+", p):
            p = f"{p}-WT"
    return p


def format_ticket_no(
    prefix: str,
    seq: int,
    width: int = DEFAULT_TICKET_SEQ_WIDTH,
) -> str:
    """生成唯一问题单编号：{prefix}-{seq:0widthd}

    例：format_ticket_no("DFKS112-WT", 1) -> "DFKS112-WT-01"
    """
    p = normalize_ticket_prefix(prefix)
    if not p:
        return ""
    n = max(1, int(seq))
    w = max(1, int(width or DEFAULT_TICKET_SEQ_WIDTH))
    return f"{p}-{n:0{w}d}"


def ensure_ticket_no(
    ticket_no: str,
    seq: int,
    prefix: str = "",
    width: int = DEFAULT_TICKET_SEQ_WIDTH,
) -> str:
    """若台账已写完整单号则沿用；否则用前缀+序号生成。

    完整单号判定：已含「-数字」后缀（如 DFKS112-WT-01）。
    若只写了前缀 DFKS112-WT，则补成 DFKS112-WT-01。
    """
    raw = (ticket_no or "").strip()
    p = normalize_ticket_prefix(prefix)

    if raw:
        # 已是 前缀-序号 形态
        if re.search(r"-\d+$", raw):
            return raw
        # 台账里只写了前缀（无流水号）
        raw_norm = normalize_ticket_prefix(raw)
        if raw_norm:
            return format_ticket_no(raw_norm, seq, width=width)
        return raw

    if p:
        return format_ticket_no(p, seq, width=width)
    return ""


def apply_tickets_to_changes(
    changes: Sequence[dict],
    tickets: Mapping[int, Ticket],
    problem_start: int = 1,
    ticket_prefix: str = "",
    ticket_seq_width: int = DEFAULT_TICKET_SEQ_WIDTH,
) -> List[dict]:
    """按问题序号把问题单信息写入每条 change。

    - 台账有该序号：用台账的「问题」描述；单号缺省或仅前缀时用 ticket_prefix 生成 DFKS112-WT-01
    - 台账无该序号但给了 ticket_prefix：仍自动生成唯一单号（便于一次版本全自动编号）
    """
    start = max(1, int(problem_start or 1))
    prefix = normalize_ticket_prefix(ticket_prefix)
    width = max(1, int(ticket_seq_width or DEFAULT_TICKET_SEQ_WIDTH))
    out: List[dict] = []
    for offset, ch in enumerate(changes or []):
        row = dict(ch)
        seq = start + offset
        row["problem_index"] = seq
        t = tickets.get(seq)
        if t is not None:
            title = t.display_title()
            no = ensure_ticket_no(t.display_no(), seq, prefix=prefix, width=width)
            row["ticket_no"] = no
            row["ticket_title"] = title
            row["ticket_seq"] = t.seq
        else:
            no = format_ticket_no(prefix, seq, width=width) if prefix else ""
            row["ticket_no"] = no
            row.setdefault("ticket_title", "")
            if no:
                row["ticket_seq"] = seq
        out.append(row)
    return out


def format_problem_heading(
    problem_index: int,
    change_type: str,
    key_display: str,
    seg_display: str,
    ticket_no: str = "",
) -> str:
    """更改单小节标题。

    有单号：（问题1，修改，DFKS112-WT-01）章节 - seg
    无单号：（问题1，修改）章节 - seg
    """
    tno = (ticket_no or "").strip()
    if tno:
        head = f"（问题{problem_index}，{change_type}，{tno}）"
    else:
        head = f"（问题{problem_index}，{change_type}）"
    key_display = key_display or ""
    seg_display = seg_display or ""
    if key_display and seg_display:
        return f"{head}{key_display} - {seg_display}"
    return f"{head}{key_display or seg_display}"


def write_ticket_template(
    path: str,
    rows: Optional[Sequence[Union[Ticket, Mapping[str, Any]]]] = None,
    n_blank: int = 10,
    ticket_prefix: str = "DFKS112-WT",
    ticket_seq_width: int = DEFAULT_TICKET_SEQ_WIDTH,
) -> str:
    """写出问题单模板（.json / .csv / .xlsx）。

    默认示例编号：DFKS112-WT-01、DFKS112-WT-02（项目型号-WT-流水号）。
    """
    path = os.path.abspath(path)
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    prefix = normalize_ticket_prefix(ticket_prefix) or "DFKS112-WT"
    width = max(1, int(ticket_seq_width or DEFAULT_TICKET_SEQ_WIDTH))

    samples: List[Ticket] = []
    if rows:
        for r in rows:
            if isinstance(r, Ticket):
                samples.append(r)
            else:
                seq = _parse_seq(r.get("seq", r.get("序号", 0))) or 0
                if seq <= 0:
                    continue
                samples.append(
                    Ticket(
                        seq=seq,
                        title=str(r.get("title") or r.get("问题") or ""),
                        ticket_no=str(r.get("ticket_no") or r.get("问题单编号") or ""),
                    )
                )
    if not samples:
        samples = [
            Ticket(
                seq=1,
                title="xxx需求变更",
                ticket_no=format_ticket_no(prefix, 1, width=width),
            ),
            Ticket(
                seq=2,
                title="xxx函数冗余",
                ticket_no=format_ticket_no(prefix, 2, width=width),
            ),
        ]
        for i in range(3, n_blank + 1):
            samples.append(
                Ticket(
                    seq=i,
                    title="",
                    ticket_no=format_ticket_no(prefix, i, width=width),
                )
            )

    ext = os.path.splitext(path)[1].lower()
    if ext == ".json":
        payload = {
            "说明": (
                "序号与更改单「问题N」对齐；问题=描述；"
                "问题单编号在一次版本变更内唯一，格式为 项目型号-WT-两位序号，"
                f"例如 {format_ticket_no(prefix, 1, width=width)}、"
                f"{format_ticket_no(prefix, 2, width=width)}"
            ),
            "ticket_prefix": prefix,
            "tickets": [
                {"seq": t.seq, "title": t.title, "ticket_no": t.ticket_no}
                for t in samples
                if t.title or t.ticket_no or t.seq <= 2
            ],
        }
        if not rows:
            payload["tickets"] = [
                {
                    "seq": 1,
                    "title": "xxx需求变更",
                    "ticket_no": format_ticket_no(prefix, 1, width=width),
                },
                {
                    "seq": 2,
                    "title": "xxx函数冗余",
                    "ticket_no": format_ticket_no(prefix, 2, width=width),
                },
            ]
        with open(path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        return path

    if ext in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import Workbook  # type: ignore
        except ImportError as exc:
            raise ImportError(
                "写出 .xlsx 需要 openpyxl。请 pip install openpyxl，或改用 .csv / .json。"
            ) from exc
        wb = Workbook()
        ws = wb.active
        ws.title = "问题单"
        ws.append(["序号", "问题", "问题单编号"])
        for t in samples:
            if not rows and t.seq > 2 and not t.title and not t.ticket_no:
                continue
            ws.append([t.seq, t.title, t.ticket_no])
        if not rows:
            for i in range(3, n_blank + 1):
                ws.append([i, "", format_ticket_no(prefix, i, width=width)])
        wb.save(path)
        return path

    # 默认 CSV
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["序号", "问题", "问题单编号"])
        if not rows:
            w.writerow([1, "xxx需求变更", format_ticket_no(prefix, 1, width=width)])
            w.writerow([2, "xxx函数冗余", format_ticket_no(prefix, 2, width=width)])
            for i in range(3, n_blank + 1):
                w.writerow([i, "", format_ticket_no(prefix, i, width=width)])
        else:
            for t in samples:
                w.writerow([t.seq, t.title, t.ticket_no])
    return path
